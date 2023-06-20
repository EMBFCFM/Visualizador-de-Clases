import openpyxl

def comparar_archivos(archivo_generado, archivo_referencia):
    # Cargar el archivo generado y el archivo de referencia
    libro_generado = openpyxl.load_workbook(archivo_generado)
    libro_referencia = openpyxl.load_workbook(archivo_referencia)

    # Obtener la hoja activa de cada libro
    hoja_generado = libro_generado.active
    hoja_referencia = libro_referencia.active

    # Obtener las columnas A y B de cada hoja
    columna_a_generado = hoja_generado['A']
    columna_b_referencia = hoja_referencia['B']

    # Crear un conjunto con los valores de la columna B del archivo de referencia
    valores_referencia = set(celda.value for celda in columna_b_referencia if celda.value is not None)

    # Comparar los valores de la columna A del archivo generado con los valores de referencia
    for celda_a_generado in columna_a_generado:
        valor_a_generado = celda_a_generado.value
        if valor_a_generado is not None and "FOLIO" not in str(valor_a_generado):
            if valor_a_generado in valores_referencia:
                celda_b_referencia = hoja_referencia.cell(column=2, row=celda_a_generado.row)
                print("Valor de la celda de la columna A:", valor_a_generado)
                print("Encontrado en:", celda_b_referencia.coordinate)
                print()

    # Cerrar los libros de Excel
    libro_generado.close()
    libro_referencia.close()

# Ejemplo de uso
archivo_generado = r'C:\Users\Usuario\Desktop\Horarios EJ 2023\Archivo.xlsx'
archivo_referencia = r'C:\Users\Usuario\Desktop\Horarios EJ 2023\6ReporteHorariosExcel.1673049882968.xlsx'

comparar_archivos(archivo_generado, archivo_referencia)
