import openpyxl

# Crear un nuevo libro de Excel
libro_excel = openpyxl.Workbook()

# Obtener la hoja activa del libro (por defecto es la primera hoja)
hoja_activa = libro_excel.active

# Leer el archivo de texto
with open(r'C:\Users\Usuario\Desktop\Horarios EJ 2023\20230616 Horario 401.txt', 'r', encoding='utf-8') as archivo_txt:
    lineas = archivo_txt.readlines()

# Variable de bandera para indicar si ya se copió la línea "SEMESTRE:"
copiado_semestre = False
encontrada_universidad = False
contador_filas = 0

# Iterar sobre las líneas del archivo de texto
for indice, linea in enumerate(lineas):
    # Eliminar los espacios en blanco al principio y al final de la línea
    linea = linea.strip()
    
    if "UNIVERSIDAD AUTONOMA DE NUEVO LEON" in linea:
        encontrada_universidad = True
        contador_filas = 0
        ultima_fila_copiada = 0

    if encontrada_universidad:
        if contador_filas < 5:
            hoja_activa.cell(row=indice+1, column=5).value = linea
            contador_filas += 1
            ultima_fila_copiada = indice + 1
        elif indice + 1 == ultima_fila_copiada + 3:
            
            
            if linea and linea[0].isdigit():
                hoja_activa.cell(row=indice+1,column=1).value = "Si, hay un digito"

    if "SEMESTRE:" in linea:
        indice_separador = linea.index("SEMESTRE:")
        caracteres = linea[indice_separador+10:]
        hoja_activa.cell(row=indice+1, column=2).value = "SEMESTRE: " + caracteres

    if "GRUPO:" in linea:
        indice_separador2 = linea.index("GRUPO:")
        car = linea[indice_separador+7:]
        hoja_activa.cell(row=indice+1, column=2).value = "GRUPO: " + car

        # Copiamos la siguiente línea en la columna 3
        siguiente_linea = lineas[indice+1].strip()
        partes = siguiente_linea.split()
        for i, parte in enumerate(partes):
            hoja_activa.cell(row=indice+2, column=i+1).value = parte
        
        #pasamos 2 filas y aplicamos el condicional
        siguiente_linea_condicional = lineas[indice+3].strip()
        while siguiente_linea_condicional and siguiente_linea_condicional != '':
            if siguiente_linea_condicional[0].isdigit():
                hoja_activa.cell(row=indice+4,column=1).value = siguiente_linea_condicional[:3]
                hoja_activa.cell(row=indice+4,column=2).value = siguiente_linea_condicional[3:12]
                hoja_activa.cell(row=indice+4,column=3).value = siguiente_linea_condicional[12:45]
                hoja_activa.cell(row=indice+4,column=4).value = siguiente_linea_condicional[45:47]
                hoja_activa.cell(row=indice+4,column=5).value = siguiente_linea_condicional[55:56]
                hoja_activa.cell(row=indice+4,column=6).value = siguiente_linea_condicional[59:65]
                hoja_activa.cell(row=indice+4,column=7).value = siguiente_linea_condicional[66:99]
                hoja_activa.cell(row=indice+4,column=8).value = siguiente_linea_condicional[100:]
           
            indice+=1
            siguiente_linea_condicional = lineas[indice+3].strip()
print('Archivo Guardado')
# Guardar el libro de Excel
libro_excel.save('pruebacopiado.xlsx')
