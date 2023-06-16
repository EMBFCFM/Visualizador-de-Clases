from tkinter import messagebox
from os import path
import os
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment
import pandas as pd


##### FUNCIONES REVISAR MINUTAS #####

def buscarErrores(archivos):
    global contarMinutas
    ordinario =      ['0','1','2','3','4','5','6','7','8','9','00','01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '34', '35', '36', '37', '38', '39', '40', '41', '42', '43', '44', '45', '46', '47', '48', '49', '50', '51', '52', '53', '54', '55', '56', '57', '58', '59', '60', '61', '62', '63', '64', '65', '66', '67', '68', '69', '70', '71', '72', '73', '74', '75', '76', '77', '78', '79', '80', '81', '82', '83', '84', '85', '86', '87', '88', '89', '90', '91', '92', '93', '94', '95', '96', '97', '98', '99', '100', 'NP','CU']
    extraordinario = ['0','1','2','3','4','5','6','7','8','9','00','01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '34', '35', '36', '37', '38', '39', '40', '41', '42', '43', '44', '45', '46', '47', '48', '49', '50', '51', '52', '53', '54', '55', '56', '57', '58', '59', '60', '61', '62', '63', '64', '65', '66', '67', '68', '69', '70', '71', '72', '73', '74', '75', '76', '77', '78', '79', '80', '81', '82', '83', '84', '85', '86', '87', '88', '89', '90', '91', '92', '93', '94', '95', '96', '97', '98', '99', '100', 'NP','CU']
    lineas = []
    for ruta in archivos:
        f = open(ruta, "r", encoding="utf8")
        leer = f.readlines()
        lineas.append(ruta)
        for line in leer:
            lineas.append(line)
        f.close()    
    datos1 = ""
    datos2 = ""
    cache1=""
    cache2=""

    profesores = []
        
    for line in reversed(lineas):
        if " FACULTAD " in line: 
            datos1 = line.lstrip()
            continue
            
        elif " CLASE " in line or " ASESORIA " in line or " 3ra. " in line:
            datos2 = line.lstrip()
            continue
            
        else:
            linea = line.lstrip()
            if len(linea) > 3:
                if linea[0].isdigit() == True:
                    if " " not in line[9:15]:
                        profesor = linea
                    check = linea[56:60]
                    if check != "":
                        if ((int(linea[0])%2) == 1 and check.rstrip() not in ordinario) or ((int(linea[0])%2) == 0 and check.rstrip() not in extraordinario) :     
                            while len(line) < 95:
                                line = line + " "
                            profesores.append(line+" "+profesor)
                            # print(linea[56:60])
                            # print(line)
                        

    datos1 = ""
    datos2 = ""
    cache1=""
    cache2=""            
    reporte = []
    nombresArchivo = []
    for line in lineas:
        if ".txt" in line:
            archivo = line
            archivo = archivo.split("/")
            archivo = archivo[-1]
        if " FACULTAD " in line: 
            datos1 = line.lstrip()
            continue
            
        elif " CLASE " in line or " ASESORIA " in line or " 3ra. " in line:
            datos2 = line.lstrip()
            continue
            
        else:
            linea = line.lstrip()
            if len(linea) > 3:
                if linea[0].isdigit() == True:
                    if " " not in line[9:15]:
                        profesor = linea
                    check = linea[56:60]
                    if check != "":
                        if ((int(linea[0])%2) == 1 and check.rstrip() not in ordinario) or ((int(linea[0])%2) == 0 and check.rstrip() not in extraordinario) :
                            if " servicio social " in datos2.lower() or " prácticas profesionales " in datos2.lower() or " practicas profesionales " in datos2.lower():
                                continue
                            if datos1 != cache1:
                                reporte.append(datos1)
                                nombresArchivo.append(archivo)
                                cache1 = datos1
                                #print(datos1)
                            if datos2 != cache2:
                                reporte.append(datos2)
                                cache2 = datos2
                                #print(datos2)
                            while len(line) < 95:
                                line = line + " "
                            reporte.append(line)
                            # print(linea[56:60])
                            # print(line)
                        
    i = 0
    for line in reporte:
        for linea in profesores:
            if line in linea:
                reporte[i] = linea 
        i+=1

    minutas = []
    for minuta in lineas:
        if " CLASE " in minuta or " ASESORIA " in minuta or " 3ra. " in minuta:
            if minuta not in minutas:
                minutas.append(minuta)

    contarMinutas = len(minutas)
    print(reporte[0:3])
    print(contarMinutas)
    print("Se analizaron:", contarMinutas, "minutas")
    crearExcelMinutas(reporte, nombresArchivo)
    #return contarMinutas

def crearExcelMinutas(reporte, nombresArchivo):
    #reporte = reversed(reporte)
    wb = Workbook()
    print("AQUI ESTA" , nombreArchivo)
    wb.save(nombreArchivo)
    sheet = wb.active
    sheet.title = "Minutas"
    i=1
    sheet.column_dimensions["A"].width = 4
    sheet.column_dimensions["B"].width = 9
    sheet.column_dimensions["C"].width = 38
    sheet.column_dimensions["D"].width = 6
    sheet.column_dimensions["E"].width = 20
    sheet.column_dimensions["F"].width = 50
    sheet.column_dimensions["G"].width = 28
    thin = Side(border_style="thin", color="000000")
    x=0
    for line in reporte:
        if " FACULTAD " in line: 
            sheet["A"+str(i)] = line.lstrip()
            sheet["A"+str(i)].border = Border(top=thin, left=thin, right=thin, bottom=thin)    
            sheet.merge_cells("A"+str(i)+":F"+str(i))
            sheet["A"+str(i)].fill = PatternFill("solid", start_color="ADADAD")
            sheet["G"+str(i)] = nombresArchivo[x]
            x+=1
        
        elif " CLASE " in line:
            sheet["A"+str(i)] = line.lstrip()
            sheet["A"+str(i)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet.merge_cells("A"+str(i)+":F"+str(i))
            sheet["A"+str(i)].fill = PatternFill("solid", start_color="C6C6C6")
            
        else:  
            sheet["A"+str(i)]= line[3]
            sheet["B"+str(i)]= line[5:15]
            sheet["C"+str(i)]= line[15:59]
            sheet["D"+str(i)]= line[59:67]
            sheet["E"+str(i)]= line[67:95]
            sheet["F"+str(i)]= line[95:]
            sheet["A"+str(i)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet["B"+str(i)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet["C"+str(i)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet["D"+str(i)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet["E"+str(i)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            sheet["F"+str(i)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        i+=1
    wb.save(nombreArchivo)

def llamarFuncionesMinutas(archivos, rutaGuardado, nombreExcel):
    global nombreArchivo 
    try:
        print("archivos:", archivos)
        if archivos == "":
            messagebox.showinfo(message = "Por favor, elija los archivos.", title = "Aviso")
            return
    except:
        messagebox.showinfo(message = "Por favor, elija los archivos.", title = "Aviso")
        return
    try:
        print("ruta", rutaGuardado)
        if rutaGuardado == "":
            messagebox.showinfo(message = "Por favor, elija la ruta a guardar.", title = "Aviso")
            return
    except:
        messagebox.showinfo(message = "Por favor, elija la ruta a guardar.", title = "Aviso")
        return
    
    if str(nombreExcel) == "":
        messagebox.showinfo(message = "Por favor, ingrese un nombre para el archivo.", title = "Aviso")
        return
    else:
        nombreArchivo= str(rutaGuardado)+"/"+str(nombreExcel) + ".xlsx"
        if path.exists(nombreArchivo) == True:
            messagebox.showinfo(message = "Ya existe un archivo con ese nombre, ingrese uno distinto.", title = "Aviso")
            return
        else:
            #try:
                buscarErrores(archivos)
                messagebox.showinfo(message = "           Excel creado.\n\nSe analizaron "+str(contarMinutas)+" minutas.", title = "Aviso", )
                print(nombreArchivo)
            #except: 
             #   messagebox.showinfo(message = "Algo salió mal, verifique sus entradas.", title = "Aviso")



##### FUNCIONES CORREOS ALUMNOS #####
def eliminarEncabezadoCorreo(archivos):
    
    for ruta in archivos:
        eliminar=["UNIVERSIDAD", "FACULTAD", "LISTADO", "Email", "---"]
        licenciaturas = [" 01 "," 02 ", " 03 ", " 04 ", " 05 ", " 06 ", " 99 "]
        licenciaturasTexto = ["MATEMATICAS", "FISICA", "CIENCIAS" ,"ACTUARIA", "MULTIMEDIA", "SEGURIDAD", "INTERCAMBIO"]
        lineas=[]
        with open(ruta, "r", encoding="utf8") as txt:
            archivo = txt.readlines()
            #lineas=[]
            lic = ""
            for line in archivo:

                imprimir = True
                if "LIC." in line or "LICENCIATURA" in line or "LICENCIADO" in line:
                    #lic = line.strip("  ")
                    imprimir = False
                    i=0
                    for licen in licenciaturasTexto:
                        if str(licen) in line:
                            lic = licenciaturas[i]
                        i+=1
                        #print(lic)
                elif len(line.lstrip()) <60:
                    imprimir = False
                for condicion in eliminar:
                    if condicion in line:
                        imprimir = False
                if imprimir == True:
                    lineas.append(lic.lstrip()+line)
        plan = ruta.split("/")
        plan = plan[-1]
        plan = plan[:-4]
        crearExcelCorreo(lineas, plan)

        # f = open("Prueba.txt", "w", encoding="utf8")

        # for linea in lineas:
        #     #print(len(linea) , linea)
        #     f.write(linea)
        # f.close()
        # print("Son", len(lineas), "lineas")

def crearExcelCorreo(lineas, plan):
    #nombreArchivo = str(rutaGuardado)+"/"+str(nombreExcel.get()) + ".xlsx"
    if path.exists(nombreArchivo) == False:
            #oportunidad = "4ta"
            wb = Workbook()
            wb.save(nombreArchivo)
    else:
        wb = load_workbook(nombreArchivo)

    wb.create_sheet(plan)
    #wb.active = 
    sheet = wb[plan]
    sheet.column_dimensions["A"].width = 4
    sheet.column_dimensions["B"].width = 10
    sheet.column_dimensions["C"].width = 42
    sheet.column_dimensions["D"].width = 5
    sheet.column_dimensions["E"].width = 40
    sheet.column_dimensions["F"].width = 13
    highlight = NamedStyle(name= "highlight")
    bd = Side(style='thin', color="000000")
    highlight.fill = PatternFill("solid", fgColor="00C0C0C0")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    highlight.alignment = Alignment(horizontal="left")
    bordes = NamedStyle(name="bordes")
    bordes.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    bordes.alignment = Alignment(horizontal="left")
    #sheet.title = "Correos"
    try:
        del wb["Sheet"]
    except:
        pass
    try:
        wb.add_named_style(bordes)
        wb.add_named_style(highlight)
    except:
        pass
    sheet["A1"].style = "highlight"
    sheet["B1"].style = "highlight"
    sheet["C1"].style = "highlight"
    sheet["D1"].style = "highlight"
    sheet["E1"].style = "highlight"
    sheet["F1"].style = "highlight"
    sheet["A1"] = "Lic."
    sheet["B1"] = "Matrícula"
    sheet["C1"] = "Nombre"
    sheet["D1"] = "Sem."
    sheet["E1"] = "correo"
    sheet["F1"] = "Teléfono"

    i=2
    while i <= len(lineas)+1:
        linea = lineas[i-2]
        #print(linea)
        separado = linea.split()
        posicion = 0
        contarInt = 0
        nombre = ""
        semestre = ""
        correo = ""
        telefono = ""
        for parte in separado:
            try:
                int(parte)
                if contarInt == 3:
                    semestre += parte
                contarInt += 1
                posicion+=1
            except:
                if "@" not in parte:
                    if any(char.isdigit() for char in parte) == False:
                        nombre += parte + " "
                        posicion+=1
                    else:
                        nombre += parte[:-2] + ""
                        semestre = parte[-1]
                        posicion+=1
                else:
                    correo += parte
                    posicion+=1
                    break
        if posicion == len(separado):
            pass
        else:
            while posicion < len(separado):
                telefono += separado[posicion]+" "
                posicion+=1
                
        print(separado)
        print(nombre)
        sheet["A"+str(i)] = int(separado[0])
        sheet["B"+str(i)] = int(separado[2])
        sheet["C"+str(i)] = nombre.rstrip()
        sheet["D"+str(i)] = int(semestre)
        sheet["E"+str(i)] = correo.rstrip()
        try:
            sheet["F"+str(i)] = int(telefono.rstrip())
        except:
            sheet["F"+str(i)] = telefono.rstrip()
        sheet["A"+str(i)].style = "bordes"
        sheet["B"+str(i)].style = "bordes"
        sheet["C"+str(i)].style = "bordes"
        sheet["D"+str(i)].style = "bordes"
        sheet["E"+str(i)].style = "bordes"
        sheet["F"+str(i)].style = "bordes"
        i+=1

    wb.save(nombreArchivo)

def llamarFuncionesCorreo(archivos, rutaGuardado, nombreExcel):
    global nombreArchivo 
    try:
        print("archivos:", archivos)
        if archivos == "":
            messagebox.showinfo(message = "Por favor, elija los archivos.", title = "Aviso")
            return
    except:
        messagebox.showinfo(message = "Por favor, elija los archivos.", title = "Aviso")
        return
    try:
        print("ruta", rutaGuardado)
        if rutaGuardado == "":
            messagebox.showinfo(message = "Por favor, elija la ruta a guardar.", title = "Aviso")
            return
    except:
        messagebox.showinfo(message = "Por favor, elija la ruta a guardar.", title = "Aviso")
        return
    
    if str(nombreExcel) == "":
        messagebox.showinfo(message = "Por favor, ingrese un nombre para el archivo.", title = "Aviso")
        return
    else:
        nombreArchivo= str(rutaGuardado)+"/"+str(nombreExcel) + ".xlsx"
        if path.exists(nombreArchivo) == True:
            messagebox.showinfo(message = "Ya existe un archivo con ese nombre, ingrese uno distinto.", title = "Aviso")
            return
        else:
            try:
                eliminarEncabezadoCorreo(archivos)
                messagebox.showinfo(message = "Excel creado.", title = "Aviso")
            except: 
                messagebox.showinfo(message = "Algo salió mal, verifique sus entradas.", title = "Aviso")


##### FUNCIONES ALUMNOS REPROBADOS #####

def eliminarEncabezadoReprobados(ruta):
    eliminar=["UNIVERSIDAD", "FACULTAD", "REPORTE", "PLAN DE ESTUDIO", "PERIODO:", "Cve", "---"]
    oportunidades = ["SEGUNDA", "CUARTA", "SEXTA"]
    with open(ruta, "r", encoding="utf8") as txt:     
        archivo = txt.readlines()
        lineas=[]
        for line in archivo:
            #print(line)
            for op in oportunidades:
                if op in line and eliminar[2] in line:
                    oportunidad = op
            for condicion in eliminar:
                #print(condicion)
                if "PLAN DE ESTUDIO:" in line:
                    planEstudio = line.split("PLAN DE ESTUDIO:", maxsplit=1)[-1].split(maxsplit=1)[0]
                if condicion in line:
                    imprimir = False
                    break
                else:
                    imprimir = True
            if imprimir == True:
                lineas.append(line)
    return lineas, planEstudio, oportunidad

def eliminarLineasReprobados(lineas):
    licenciaturas = [" 01 "," 02 ", " 03 ", " 04 ", " 05 ", " 06 ", " 99 "]
    licExcel = []
    licEx = 0
    for line in lineas:
        for lic in licenciaturas:
            if lic in line:
                licEx = lic
                print(line)
        licExcel.append(licEx)

    i=1
    indices = []
    for lic in licExcel:
        if lic != licExcel[i]:
            indices.append(i-1)
        i+=1
        if i == len(licExcel):
            break

    for ind in reversed(indices):
        licExcel.pop(ind)
        lineas.pop(ind)
        licExcel.pop(ind)
        lineas.pop(ind)

    i=0
    indices = []
    for line in lineas:
        if len(line.split()) < 3:
            indices.append(i)  
        i+=1

    for ind in indices:
        licExcel.pop(ind)
        lineas.pop(ind)
    
    return licExcel, lineas

def crearListasReprobados(lineas):
    nombres = []
    matriculas = []
    materias = []
    cves = []
    cont=0
    for line in lineas:
        nombre = []
        matricula = []
        materia = []
        cve = []
        x = line.split()
        for i in x:
            if i.isdigit():
                if len(nombre) == 0 and len(i) > 3:
                    matricula.append(i)
                elif len(cve) == 0 and len(materia) == 0:
                    cve.append(i)
            elif len(i) == 3 and i[1].isdigit() and len(cve) == 0:
                cve.append(i)
            elif len(cve) == 1:
                materia.append(i)
            elif len(matricula) == 2 and len(cve) == 0:
                nombre.append(i)
            
            # elif len(matricula) == 1:       #Condición segunda materia alumno
            #     cve.append(matricula[0])
            #     matricula.clear()
            #     materia.append(i)
                
        if len(matricula) == 0 and len(nombre) == 0:
            matricula = matriculas[-1]
            nombre = nombres[-1]
            # print("materia",materia)
            # print("matricula",matricula)
            # print("nombre",nombre)
            # print("cve" , cve)
            matriculas.append(matricula)
            nombres.append(nombre)
            cves.append(cve[0])
            materias.append(' '.join(materia))
            
        else:
            # print("materia",materia)
            # print("matricula",matricula)
            # print("nombre",nombre)
            # print("cve" , cve)
            nombres.append(' '.join(nombre))
            materias.append(' '.join(materia))
            cves.append(cve[0])
            matriculas.append(matricula[1])
            
        #print(matriculas[cont], "", nombres[cont], "", cves[cont], "", materias[cont])
        cont +=1
    return nombres, materias, cves, matriculas


def crearExcelReprobados(nombres, materias, cves, matriculas, planEstudio, licExcel, nombreArchivo, oportunidad):
    if oportunidad == "SEGUNDA":
        oportunidad = "2da"
        opHoja = "Segundas"
    elif oportunidad == "CUARTA":
        oportunidad = "4ta"
        opHoja = "Cuartas"
    elif oportunidad == "SEXTA":
        oportunidad = "6ta"
        opHoja = "Sextas"

    if path.exists(nombreArchivo) == False:
        #oportunidad = "4ta"
        wb = Workbook()
        wb.save(nombreArchivo)

    wb = load_workbook(filename=nombreArchivo)
    if opHoja not in wb.sheetnames:
        if len(wb.sheetnames) == 1 and wb.active.title == "Sheet": 
            sheet = wb.active
            sheet.title = opHoja
        else:
            if opHoja == "Segundas":
                sheet = wb.create_sheet(title=opHoja, index=0)
            elif opHoja == "Cuartas":
                sheet = wb.create_sheet(title=opHoja, index=1)
            elif opHoja == "Sextas":
                sheet = wb.create_sheet(title=opHoja, index=2)
                #sheet.title = opHoja
        sheet.column_dimensions["A"].width = 5
        sheet.column_dimensions["B"].width = 12
        sheet.column_dimensions["C"].width = 12
        sheet.column_dimensions["D"].width = 9
        sheet.column_dimensions["E"].width = 40
        sheet.column_dimensions["F"].width = 5
        sheet.column_dimensions["G"].width = 42
        highlight = NamedStyle(name= "highlight")
        bd = Side(style='thin', color="000000")
        highlight.fill = PatternFill("solid", fgColor="00C0C0C0")
        highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        highlight.alignment = Alignment(horizontal="center")
        try:
            wb.add_named_style(highlight)
        except:
            print("Style already exists")
        sheet["A1"].style = "highlight"
        sheet["B1"].style = "highlight"
        sheet["C1"].style = "highlight"
        sheet["D1"].style = "highlight"
        sheet["E1"].style = "highlight"
        sheet["F1"].style = "highlight"
        sheet["G1"].style = "highlight"
        sheet["A1"] = "Plan"
        sheet["B1"] = "Oportunidad"
        sheet["C1"] = "Licenciatura"
        sheet["D1"] = "Matrícula"
        sheet["E1"] = "Nombre"
        sheet["F1"] = "Cve"
        sheet["G1"] = "Materia"
        sheet["E1"].alignment = Alignment(horizontal="left")
        sheet["G1"].alignment = Alignment(horizontal="left")
        i=2
        print("nombres ", len(nombres))

        while i <= len(nombres)+1:
            #print(i)
            sheet["A"+str(i)] = int(planEstudio)
            sheet["B"+str(i)] = oportunidad
            sheet["C"+str(i)] = str(licExcel[i-2])
            sheet["D"+str(i)] = int(matriculas[i-2])
            sheet["E"+str(i)] = nombres[i-2]
            sheet["F"+str(i)] = str(cves[i-2])
            sheet["G"+str(i)] = materias[i-2]
            sheet["A"+str(i)].alignment = Alignment(horizontal="center")
            sheet["B"+str(i)].alignment = Alignment(horizontal="center")
            sheet["C"+str(i)].alignment = Alignment(horizontal="center")
            sheet["D"+str(i)].alignment = Alignment(horizontal="center")
            sheet["F"+str(i)].alignment = Alignment(horizontal="center")
            i+=1
            #print(i)
        wb.save(nombreArchivo)
    else:
        wb = load_workbook(filename=nombreArchivo)
        sheet = wb[opHoja]
        #oportunidad = "4ta"
        j = sheet.max_row+1
        i=0
        while i < len(nombres):
            #print(i)
            sheet["A"+str(j)] = int(planEstudio)
            sheet["B"+str(j)] = oportunidad
            sheet["C"+str(j)] = str(licExcel[i])
            sheet["D"+str(j)] = int(matriculas[i])
            sheet["E"+str(j)] = nombres[i]
            sheet["F"+str(j)] = str(cves[i])
            sheet["G"+str(j)] = materias[i]
            sheet["A"+str(j)].alignment = Alignment(horizontal="center")
            sheet["B"+str(j)].alignment = Alignment(horizontal="center")
            sheet["C"+str(j)].alignment = Alignment(horizontal="center")
            sheet["D"+str(j)].alignment = Alignment(horizontal="center")
            sheet["F"+str(j)].alignment = Alignment(horizontal="center")
            i+=1
            j+=1
            #print(i)
        wb.save(nombreArchivo)
        
def parametrosExcelReprobados(archivos, rutaGuardado, nombreExcel):
    x=0
    while x==0:
        try:
            nombreArchivo = str(rutaGuardado)+"/"+str(nombreExcel) + ".xlsx"
            x=1
        except:
            messagebox.showinfo(message="Por favor seleccione la ruta a guardar el archivo", title="Aviso")
            #rutaGuardar()

    print(nombreArchivo)
    if nombreArchivo == ".xlsx":
        messagebox.showwarning(message = "Por favor, ingrese un nombre válido", title = "Aviso")
        return
    if path.exists(nombreArchivo) == True:
        messagebox.showinfo(message = "Ya existe un archivo con ese nombre", title = "Aviso")
        return
    try:
        for ruta in archivos:
            prim = eliminarEncabezadoReprobados(ruta)
            planEstudio = prim[1]
            lineas = prim[0]
            oportunidad = prim[2]

            seg = eliminarLineasReprobados(lineas)
            licExcel = seg[0]
            lineas = seg[1]

            terc = crearListasReprobados(lineas)
            nombres = terc[0]
            materias = terc[1]
            cves = terc[2]
            matriculas = terc[3]

            crearExcelReprobados(nombres, materias, cves, matriculas, planEstudio, licExcel, nombreArchivo, oportunidad)
    except:
        messagebox.showwarning(message = "Algo salió mal, verifique las entradas.", title = "Aviso")
        return
    if path.exists(nombreArchivo) == False:
        #messagebox.showwarning(message = "Por favor, seleccione los archivos", title = "Aviso")
        return
    else:
        messagebox.showinfo(message = "Excel creado", title = "Aviso")

def reportesAlumnos(archivos,opcion,opo,clave):
    #x = plan.get()
    y = 0
    # Obtenemos en nombre del archivo y eliminamos la etiqueta de .txt
    name = os.path.basename(archivos)
    name = name.removesuffix('.txt')
    # Obtenemos la fecha de creacion del archivo
    creation = os.path.getctime(archivos)
    creation = datetime.datetime.fromtimestamp(creation)
    # Sacamos el año y el mes y se transforman en strings
    year = creation.strftime("%Y")
    month = creation.strftime("%m")
    # Se hace la comprobacion de que se ha seleccionado un archivo
    if name == '':
        #nofile()
        print('ERROR')
    else:
        # Header
        header1 = "UNIVERSIDAD AUTONOMA DE NUEVO LEON"
        header2 = "FACULTAD DE CIENCIAS FÍSICO MATEMÁTICAS"
        if "ago" in name:
            header4 = "PERIODO: Agosto-Diciembre "+year
        elif "ene" in name:
            header4 = "PERIODO: Enero-Junio "+year
        elif "01" == month or "02" == month or "03" == month or "04" == month or "05" == month or "06" == month:
            header4 = "PERIODO: Enero-Junio "+year
        elif "08" == month or "09" == month or "10" == month or "11" == month or "12" == month or "07" == month:
            header4 = "PERIODO: Agosto-Diciembre "+year
        else:
            #wrongopt()
            messagebox.showinfo(message = "Por favor, elija los archivos.", title = "Aviso")
            return
        # Seleccion del tipo de formato segun el tipo de lista requerida
        if "401" in name and "Alumnos con escuela de procedencia" in opcion:
            header5 = "PLAN DE ESTUDIO: 401 MODELO ACADEMICO V1"
            header3 = "ALUMNOS DEL TODOS"
            y = 1
        elif "420" in name and "Alumnos con escuela de procedencia" in opcion:
            header5 = "420 Mod. Edu V2 / Mod. Acad. V3"
            header3 = "ALUMNOS DEL TODOS"
            y = 1
        elif "430" in name and "Alumnos con escuela de procedencia" in opcion:
            header5 = "430 Modelo Academico 2020"
            header3 = "ALUMNOS DEL TODOS"
            y = 1
        elif "440" in name and "Alumnos con escuela de procedencia" in opcion:
            header5 = "PLAN DE ESTUDIO 440 MODELO ACADEMICO 2022"
            header3 = "ALUMNOS DEL TODOS"
            y = 1
        elif "Alumnos" in name and "Lista de Alumnos" in opcion:
            header3 = "LISTADO DE ALUMNOS"
            y = 2
        if "mod1" in name and "Lista de Alumnos" in opcion:
            header3 = "LISTADO DE ALUMNOS"
            mod = "1"
            y = 3
        elif "mod4" in name and "Lista de Alumnos" in opcion:
            header3 = "LISTADO DE ALUMNOS"
            mod = "4"
            y = 3
        elif "mod5" in name and "Lista de Alumnos" in opcion:
            header3 = "LISTADO DE ALUMNOS"
            mod = "5"
            y = 3
        elif "Promedios" in name or "PROMEDIOS" in name:
            y = 4
            if "Listas con promedios" in opcion and "420" in name:
                header3 = "PROMEDIOS BASE KARDEX PLAN 420"
            elif "Listas con promedios" in opcion and "430" in name:
                if "kardex" in name:
                    header3 = "PROMEDIOS BASE KARDEX PLAN 430"
                elif "semestre" in name:
                    if "compelto" in name:
                        header3 = "PROMEDIOS BASE SEMESTRE COMPLETO PLAN 430"
                    else:
                        header3 = "PROMEDIOS BASE SEMESTRE PLAN 430"
                elif "inscripcion" in name:
                    header3 = "LISTADO DE PROMEDIOS PLAN 430"
            elif "Listas con promedios" in opcion and "440" in name:
                header3 = "PROMEDIOS BASE KARDEX PLAN 440"
            elif "BASE SEM" in name:
                if "MOD 1" in name:
                    header3 = "PROMEDIOS BASE SEMESTRE POR CARRERA DE LA MODALIDAD 1"
                elif "MOD 4" in name:
                    header3 = "PROMEDIOS BASE SEMESTRE POR CARRERA DE LA MODALIDAD 4"
                elif "MOD 5" in name:
                    header3 = "PROMEDIOS BASE SEMESTRE POR CARRERA DE LA MODALIDAD 5"
            else:
                y = 0
                messagebox.showinfo(message = "Por favor, elija los archivos.", title = "Aviso")
                #wrongopt()
                return
        elif "Directorio" in name and "Directorio de Alumnos con promedio" in opcion:
            header3 = "DIRECTORIO DE ALUMNOS"
            y = 5
        elif "Horarios" in name and "Horarios de Alumnos con promedio por Plan/Modalidad" in opcion:
            y = 6
            if "430" in name:
                header5 = "PLAN DE ESTUDIO: 430 Modelo Academico 2020"
            elif "401" in name:
                header5 = "PLAN DE ESTUDIO: 401 MODELO ACADEMICO V1"
            elif "420" in name:
                header5 = "PLAN DE ESTUDIO: 420 Mod. Edu V2 / Mod. Acad. V3"
            else:
                y = 0
                messagebox.showinfo(message = "Por favor, elija los archivos.", title = "Aviso")
                #wrongopt()
                return
        if y == 1:
            # Especificaciones de como se va a leer la informacion de las columnas del archivo de texto
            colspecs = [(0, 2), (3, 10), (11, 47), (48, 53),
                        (54, 89), (90, 91), (92, 96), (97, 99)]
            # Lectura del archivo de texto
            df = pd.read_fwf(archivos, colspecs=colspecs, header=None,
                             index_col=0, skiprows=9, dtype={3: "string", 7: "string"})
            # Nombres de las Columnas en las que se divide la tabla
            df.columns = ['MATRICULA', 'NOMBRE', ' ',
                          'PROCEDENCIA', 'S', 'CARR', 'SEM']
            # Se eliminan las filas que no son necesarias basado en la informacion de la columna S
            df = df[df["S"].str.contains(" |S|-|1|2|3|4|5|6|7|8|9") == False]
            # Crear el archivo Excel
            writer = pd.ExcelWriter(name+'.xlsx', engine='xlsxwriter')
            df.to_excel(writer, startrow=6, startcol=0,
                        engine='xlsxwriter', index=False)
            # Columnas que se van a modificar de tamaño
            col_idx_mat = df.columns.get_loc('MATRICULA')
            col_idx_nom = df.columns.get_loc('NOMBRE')
            col_idx_proc = df.columns.get_loc('PROCEDENCIA')
            # Escribir el header en el excel
            worksheet = writer.sheets['Sheet1']
            worksheet.write(0, 1, header1)
            worksheet.write(1, 1, header2)
            worksheet.write(2, 1, header3)
            worksheet.write(3, 1, header4)
            worksheet.write(4, 1, header5)
            # Formato personalizado de las columnas
            worksheet.set_column(col_idx_mat, col_idx_mat, 11)
            worksheet.set_column(col_idx_nom, col_idx_nom, 40)
            worksheet.set_column(col_idx_proc, col_idx_proc, 40)
            writer.close()
            messagebox.showinfo(message = "Excel creado.", title = "Aviso")
            ##success()
            y = 0
        elif y == 2:
            # Especificaciones de como se va a leer la informacion de las columnas del archivo de texto
            colspecs = [(0, 0), (1, 5), (6, 14), (15, 54),
                        (55, 58), (59, 96), (97, 111)]
            # Lectura del archivo de texto
            df = pd.read_fwf(archivos, colspecs=colspecs, header=None,
                             index_col=0, skiprows=9, dtype={4: "string", 6: "string"})
            # Nombres de las Columnas en las que se divide la tabla
            df.columns = ['Num.', 'Mat.', 'Nombre', 'Sem', 'Email', 'Telefono']

            # Se eliminan las filas que no son necesarias basado en la informacion de la columna Sem
            df = df[df["Sem"].str.contains(
                "TIC|CA|UTA|RIA|ION|IN|ICO|MA|VO|S|202|---") == False]
            # Insertamos la columna con el valor de la carrera de cada estudiante
            carrera = ['', 'LM', 'LF', 'LCC', 'LA', 'LMAD', 'LSTI', 'IA']
            num = list(df["Num."])
            num = [eval(i) for i in num]
            lista = []
            # Se crea la lista de las carreras a las que pertenece cada estudiante
            i = 0
            for k in range(len(num)):
                if num[k] == 1:
                    i += 1
                    lista.append(carrera[i])
                elif num != 1:
                    lista.append(carrera[i])
            # Se insterta la lista dentro de la matriz de datos en una nueva columna
            df.insert(loc=1, column="Carrera", value=lista)
            # Crear el archivo Excel
            writer = pd.ExcelWriter(name+'.xlsx', engine='xlsxwriter')
            df.to_excel(writer, startrow=6, startcol=0,
                        engine='xlsxwriter', index=False)
            # Columnas que se van a modificar de tamaño
            col_idx_mat = df.columns.get_loc('Mat.')
            col_idx_nom = df.columns.get_loc('Nombre')
            col_idx_email = df.columns.get_loc('Email')
            col_idx_tele = df.columns.get_loc('Telefono')
            # Escribir el header en el excel
            worksheet = writer.sheets['Sheet1']
            worksheet.write(1, 1, header1)
            worksheet.write(2, 1, header2)
            worksheet.write(3, 1, header3)
            worksheet.write(4, 1, header4)
            # Formato personalizado de las columnas
            worksheet.set_column(col_idx_mat, col_idx_mat, 11)
            worksheet.set_column(col_idx_nom, col_idx_nom, 40)
            worksheet.set_column(col_idx_email, col_idx_email, 40)
            worksheet.set_column(col_idx_tele, col_idx_tele, 15)
            writer.close()
            messagebox.showinfo(message = "Excel creado.", title = "Aviso")
            #success()
            y = 0
        elif y == 3:
            # Especificaciones de como se va a leer la informacion de las columnas del archivo de texto
            colspecs = [(0, 4), (5, 14), (15, 24), (25, 77),
                        (78, 82)]
            # Lectura del archivo de texto
            df = pd.read_fwf(archivos, colspecs=colspecs, header=None,
                             index_col=0, skiprows=9)
            # Nombres de las Columnas en las que se divide la tabla
            df.columns = ['Num.', 'Mat.', 'Nombre', 'Sem']

            # Se eliminan las filas que no son necesarias basado en la informacion de la columna Sem
            df = df[df["Num."].str.contains(" ") == False]
            df = df[df["Sem"].str.contains("Pag") == False]
            df = df[df["Sem"].str.contains("mest") == False]
            # Insertamos la columna con el valor de la carrera de cada estudiante
            if mod == '1':
                carrera = ['', 'LM', 'LF', 'LCC', 'LA', 'LMAD', 'LSTI', 'IA']
            elif mod == '4':
                carrera = ['', 'LM', 'LF', 'LA', 'LSTI']
            elif mod == '5':
                carrera = ['', 'LCC']
            num = list(df["Num."])
            sem = list(df["Sem"])
            num = [eval(i) for i in num]
            lista = []
            # Se crea la lista de las carreras a las que pertenece cada estudiante
            i = 0
            for k in range(len(num)):
                if num[k] == 1 and sem[k] == "01":
                    i += 1
                    lista.append(carrera[i])
                elif num != 1:
                    lista.append(carrera[i])
            # Se insterta la lista dentro de la matriz de datos en una nueva columna
            df.insert(loc=1, column="Carrera", value=lista)
            df.insert(loc=1, column="Mod.", value=mod)
            # Crear el archivo Excel
            writer = pd.ExcelWriter(name+'.xlsx', engine='xlsxwriter')
            df.to_excel(writer, startrow=6, startcol=0,
                        engine='xlsxwriter', index=False)
            # Columnas que se van a modificar de tamaño
            col_idx_mat = df.columns.get_loc('Mat.')
            col_idx_nom = df.columns.get_loc('Nombre')
            # Escribir el header en el excel
            worksheet = writer.sheets['Sheet1']
            worksheet.write(1, 1, header1)
            worksheet.write(2, 1, header2)
            worksheet.write(3, 1, header3)
            worksheet.write(4, 1, header4)
            # Formato personalizado de las columnas
            worksheet.set_column(col_idx_mat, col_idx_mat, 11)
            worksheet.set_column(col_idx_nom, col_idx_nom, 45)
            writer.close()
            messagebox.showinfo(message = "Excel creado.", title = "Aviso")
            #success()
            y = 0
        elif y == 4:
            # Especificaciones de como se va a leer la informacion de las columnas del archivo de texto
            if "kardex" in name:
                colspecs = [(0, 5), (6, 14), (15, 55), (56, 62),
                            (63, 70), (71, 79), (80, 85), (86, 90), (91, 95)]
            elif "semestre" in name:
                if "completo" in name:
                    colspecs = [(0, 5), (6, 14), (15, 56), (57, 60), (79, 85)]
                else:
                    colspecs = [(0, 5), (6, 14), (15, 56), (57, 60),
                                (61, 66), (67, 72), (73, 79), (80, 83), (83, 85), (86, 93)]
            elif "inscripcion" in name:
                colspecs = [(0, 5), (6, 14), (15, 56), (57, 60),
                            (61, 66), (70, 78), (79, 85)]
            elif "MOD" in name:
                colspecs = [(0, 5), (6, 14), (15, 56), (57, 60), (65, 72),
                            (61, 64), (72, 79), (80, 85)]
            # Lectura del archivo de texto
            df = pd.read_fwf(archivos, colspecs=colspecs, header=None,
                             index_col=0, skiprows=4)
            # Nombres de las Columnas en las que se divide la tabla
            if "kardex" in name:
                df.columns = ['Matricula', 'Nombre', 'Plan',
                              'Tot-Ext', 'Cla-Aprob', 'Prom', 'Sem', 'S.A']
            elif "semestre" in name:
                if "completo" in name:
                    df.columns = ['Matricula', 'Nombre', 'Plan', 'Prom']
                else:
                    df.columns = ['Matricula', 'Nombre', 'Plan',
                                  'Tot-Ext', 'Cla-Aprob', 'Prom', 'Sem', 'T', 'Aula']
            elif "inscripcion" in name:
                df.columns = ['Matricula', 'Nombre',
                              'Plan', 'Tot-Ext', 'Cla-Cur', 'Prom']
            elif "MOD" in name:
                df.columns = ['Matricula', 'Nombre', 'Plan', 'Semestre',
                              'Tot-Ext', 'Cla-Aprob', 'Prom']
            if "completo" in name:
                df = df[df["Nombre"].str.contains(
                    "Nombre|----------------------------------------|PERIODO|FACULTAD|UNIVERSIDAD|PROMEDIOS") == False]
            elif "MOD" in name:
                df = df[df["Nombre"].str.contains(
                    "Nombre|----------------------------------------|FACULTAD|UNIVERSIDAD") == False]
            else:
                df = df[df["Nombre"].str.contains(
                    "Nombre|----------------------------------------|PERIODO|FACULTAD|UNIVERSIDAD") == False]
            i = 0
            lista = []
            listaS = []
            listaN = list(df["Nombre"])
            carr = ""
            if "prom401" in x or "prom420" in x or "prom430" in x or "prom440" in x:
                j = 0
                sem = ""
                if "completo" in name:
                    for k in range(len(listaN)):
                        if "CARRERA:" not in listaN[k]:
                            if "TRE:" not in listaN[k]:
                                lista.append(carr)
                        elif "01" in listaN[k]:
                            carr = "LM"
                        elif "02" in listaN[k]:
                            carr = "LF"
                        elif "03" in listaN[k]:
                            carr = "LCC"
                        elif "04" in listaN[k]:
                            carr = "LA"
                        elif "05" in listaN[k]:
                            carr = "LMAD"
                        elif "06" in listaN[k]:
                            carr = "LSTI"
                    df = df[df["Nombre"].str.contains("CARRERA:") == False]
                    listaT = list(df["Nombre"])
                    for k in range(len(listaT)):
                        if "TRE:" not in listaT[k]:
                            listaS.append(sem)
                        elif "1" in listaT[k]:
                            sem = "1"
                        elif "2" in listaT[k]:
                            sem = "2"
                        elif "3" in listaT[k]:
                            sem = "3"
                        elif "4" in listaT[k]:
                            sem = "4"
                        elif "5" in listaT[k]:
                            sem = "5"
                        elif "6" in listaT[k]:
                            sem = "6"
                        elif "7" in listaT[k]:
                            sem = "7"
                        elif "8" in listaT[k]:
                            sem = "8"
                        elif "9" in listaT[k]:
                            sem = "9"
                    df = df[df["Nombre"].str.contains("TRE:") == False]
                    df.insert(loc=1, column="Semestre", value=listaS)
                else:
                    for i in range(len(listaN)):
                        if "CARRERA:" not in listaN[i]:
                            lista.append(carr)
                        elif "01" in listaN[i]:
                            carr = "LM"
                        elif "02" in listaN[i]:
                            carr = "LF"
                        elif "03" in listaN[i]:
                            carr = "LCC"
                        elif "04" in listaN[i]:
                            carr = "LA"
                        elif "05" in listaN[i]:
                            carr = "LMAD"
                        elif "06" in listaN[i]:
                            carr = "LSTI"
                    df = df[df["Nombre"].str.contains("CARRERA:") == False]
            elif "sembxmod" in x:
                sem = ""
                for i in range(len(listaN)):
                    if "CARRERA:" not in listaN[i]:
                        lista.append(carr)
                    elif "01" in listaN[i]:
                        carr = "LM"
                    elif "02" in listaN[i]:
                        carr = "LF"
                    elif "03" in listaN[i]:
                        carr = "LCC"
                    elif "04" in listaN[i]:
                        carr = "LA"
                    elif "05" in listaN[i]:
                        carr = "LMAD"
                    elif "06" in listaN[i]:
                        carr = "LSTI"
                df = df[df["Nombre"].str.contains("CARRERA:") == False]
                df.insert(loc=1, column="Carrera", value=lista)
                listaT = list(df["Semestre"])
                for i in range(len(listaT)):
                    if pd.isnull(listaT[i]) == True:
                        listaT[i] = "NaN"
                for i in range(len(listaT)):
                    if "NaN" in listaT[i]:
                        listaS.append(sem)
                    elif "1" in listaT[i]:
                        sem = "1"
                    elif "2" in listaT[i]:
                        sem = "2"
                    elif "3" in listaT[i]:
                        sem = "3"
                    elif "4" in listaT[i]:
                        sem = "4"
                    elif "5" in listaT[i]:
                        sem = "5"
                    elif "6" in listaT[i]:
                        sem = "6"
                    elif "7" in listaT[i]:
                        sem = "7"
                    elif "8" in listaT[i]:
                        sem = "8"
                    elif "9" in listaT[i]:
                        sem = "9"
                df["Semestre"] = listaT
                df = df[df["Matricula"].notna()]
                df["Semestre"] = listaS
            # Se insterta la lista dentro de la matriz de datos en una nueva columna
            if x != "sembxmod":
                df.insert(loc=1, column="Carrera", value=lista)
            # Crear el archivo Excel
            writer = pd.ExcelWriter(name+'.xlsx', engine='xlsxwriter')
            df.to_excel(writer, startrow=6, startcol=0,
                        engine='xlsxwriter', index=False)
            # Columnas que se van a modificar de tamaño
            col_idx_mat = df.columns.get_loc('Matricula')
            col_idx_nom = df.columns.get_loc('Nombre')
            # Escribir el header en el excel
            worksheet = writer.sheets['Sheet1']
            worksheet.write(1, 1, header1)
            worksheet.write(2, 1, header2)
            worksheet.write(3, 1, header3)
            worksheet.write(4, 1, header4)
            # Formato personalizado de las columnas
            worksheet.set_column(col_idx_mat, col_idx_mat, 11)
            worksheet.set_column(col_idx_nom, col_idx_nom, 45)
            writer.close()
            messagebox.showinfo(message = "Excel creado.", title = "Aviso")
            #success()
            y = 0
        elif y == 5:
            colspecs = [(0, 0), (0, 8), (8, 49), (49, 99),
                        (99, 140)]
            # Lectura del archivo de texto
            df = pd.read_fwf(archivos, colspecs=colspecs, header=None,
                             index_col=0, skiprows=7)
            # Nombres de las Columnas en las que se divide la tabla
            df.columns = ['Matricula', 'Nombre',
                          'Calle - Colonia', 'Ciudad - Estado']
            # Se eliminan los espacios en que pueden alterar
            # el funcionamiento del codigo al recorrer las listas
            df = df[df["Nombre"].str.contains(
                "UNIVERSIDAD AUTONOMA DE NUEVO LEON") == False]
            df = df[df["Calle - Colonia"].str.contains(
                "UNIVERSIDAD AUTONOMA DE NUEVO LEON") == False]
            df = df[df["Calle - Colonia"].str.contains(
                "Calle - Colonia") == False]
            df = df[df["Calle - Colonia"].str.contains(
                "--------------------------------------------------") == False]
            # Listas que van a contener los datos temporalmente
            tel = []
            tut = []
            prom = []
            extra = []
            mail = []
            email = []
            promF = []
            sem = []
            carr = []
            calT = list(df['Calle - Colonia'])
            nomL = list(df["Nombre"])
            ciuE = list(df['Ciudad - Estado'])
            # Ciclos para obtener los datos del Dataframe y ponerlos en una lista
            for k in range(len(nomL)):
                if "Telefono :" in nomL[k]:
                    tel.append(nomL[k])
                elif "Promedio:" in nomL[k]:
                    prom.append(nomL[k])
            for k in range(len(calT)):
                if "Tutor:" in calT[k]:
                    tut.append(calT[k])
            for k in range(len(ciuE)):
                if "e-mail:" in calT[k]:
                    if pd.isnull(ciuE[k]) == True:
                        mail.append("@NOMAIL.COM")
                    else:
                        mail.append(ciuE[k])
                    extra.append(calT[k])
            for k in range(len(tel)):
                temp = tel[k]
                temp = temp.split(':')[1]
                tel[k] = temp
            for k in range(len(tut)):
                temp = tut[k]
                temp = temp.split(':')[1]
                tut[k] = temp
            for k in range(len(prom)):
                temp = prom[k]
                temp = temp.split()[1].split()[0]
                promF.append(temp)
            for k in range(len(extra)):
                temp = extra[k]
                temp1 = extra[k]
                temp2 = extra[k]
                temp = temp.split('e-mail:')[1]
                temp1 = temp1.split('Sem:  ')[1].split(' C')[0]
                temp2 = temp2.split('Carr: ')[1].split(' e')[0]
                email.append("".join([temp, mail[k]]))
                sem.append(temp1)
                carr.append(temp2)
            # Se eliminan las filas que no son necesarias
            df = df[df["Matricula"].notna()]
            # Insertamos la columna con el valor de la carrera de cada estudiante
            carrera = ['LM', 'LF', 'LCC', 'LA', 'LMAD', 'LSTI']
            for k in range(len(carr)):
                if carr[k] == "01":
                    carr[k] = "LM"
                elif carr[k] == "02":
                    carr[k] = "LF"
                elif carr[k] == "03":
                    carr[k] = "LCC"
                elif carr[k] == "04":
                    carr[k] = "LA"
                elif carr[k] == "05":
                    carr[k] = "LMAD"
                elif carr[k] == "06":
                    carr[k] = "LSTI"
            # Se insterta la lista dentro de la matriz de datos en una nueva columna
            df.insert(loc=4, column="E-mail", value=email)
            df.insert(loc=4, column="Carrera", value=carr)
            df.insert(loc=4, column="Semestre", value=sem)
            df.insert(loc=4, column="Promedio", value=promF)
            df.insert(loc=4, column="Padre o tutor", value=tut)
            df.insert(loc=4, column="Telefono", value=tel)
            # Crear el archivo Excel
            writer = pd.ExcelWriter(name+'.xlsx', engine='xlsxwriter')
            df.to_excel(writer, startrow=6, startcol=0,
                        engine='xlsxwriter', index=False)
            # Columnas que se van a modificar de tamaño
            col_idx_mat = df.columns.get_loc('Matricula')
            col_idx_nom = df.columns.get_loc('Nombre')
            col_idx_calle = df.columns.get_loc('Calle - Colonia')
            col_idx_ciu = df.columns.get_loc('Ciudad - Estado')
            col_idx_tel = df.columns.get_loc('Telefono')
            col_idx_tut = df.columns.get_loc('Padre o tutor')
            col_idx_prom = df.columns.get_loc('Promedio')
            col_idx_mail = df.columns.get_loc('E-mail')
            # Escribir el header en el excel
            worksheet = writer.sheets['Sheet1']
            worksheet.write(1, 1, header1)
            worksheet.write(2, 1, header2)
            worksheet.write(3, 1, header3)
            worksheet.write(4, 1, header4)
            # Formato personalizado de las columnas
            worksheet.set_column(col_idx_mat, col_idx_mat, 11)
            worksheet.set_column(col_idx_tel, col_idx_tel, 15)
            worksheet.set_column(col_idx_nom, col_idx_nom, 45)
            worksheet.set_column(col_idx_tut, col_idx_tut, 40)
            worksheet.set_column(col_idx_calle, col_idx_calle, 50)
            worksheet.set_column(col_idx_ciu, col_idx_ciu, 40)
            worksheet.set_column(col_idx_prom, col_idx_prom, 10)
            worksheet.set_column(col_idx_mail, col_idx_mail, 45)
            writer.close()
            y = 0
            #success()
            messagebox.showinfo(message = "Excel creado.", title = "Aviso")
        elif y == 6:
            if opo != '' and (opo <= '0' or opo >= '7'):
                y = 0
                #wrongopt()
                return
            else:
                colspecs = [(0, 5), (5, 52), (52, 55), (55, 59),
                            (60, 65), (66, 71), (72, 78), (79, 82)]
                df = pd.read_fwf(archivos, colspecs=colspecs, header=None,
                                 index_col=0, skiprows=9)
                df.columns = ['Informacion', 'Creditos',
                              'Grupo', 'Hora', 'Frecuencia', 'Aula', 'Op']
                lista = []
                op = []
                count = 0
                A = list(df['Creditos'])
                # Se rellenan los valores vacios al inicio de la lista de creditos
                # Y se reemplaza la columna por una nueva con valores NaN
                for i in range(len(A)):
                    if pd.isnull(A[i]) == True:
                        lista.append('NaN')
                    else:
                        lista.append(A[i])
                df['Creditos'] = lista
                df = df[df['Creditos'].str.contains(
                    "D|ÍS|UM|el|--|RE|JU|du") == False]
                Lop = list(df['Op'])
                Linf = list(df['Informacion'])
                x = 0
                j = 0
                y = 0
                z = False
                df2 = pd.DataFrame()
                # Ciclo para tomar las coordenadas en el DataFrame y separarlas por bloques
                # segun se cumpla la condicion, en este caso los alumnos con al menos una 5ta
                for i in range(len(Lop)):
                    if pd.isnull(Lop[i]):
                        count += 1
                if clave == '':
                    for i in range(len(Lop)):
                        if pd.isnull(Lop[i]) == True and j == 0:
                            x = i
                        elif pd.isnull(Lop[i]) == True & j == 1:
                            y = i
                            j = 2
                        elif opo in Lop[i]:
                            z = True
                            j = 1
                        if j == 2 and z == True:
                            j = 0
                            z = False
                            for x in range(x, y):
                                op.append(x)
                                if x == y-1:
                                    x = y
                else:
                    for i in range(len(Lop)):
                        if pd.isnull(Lop[i]) == True and j == 0:
                            x = i
                        elif pd.isnull(Lop[i]) == True & j == 1:
                            y = i
                            j = 2
                        elif pd.isnull(Linf[i]) == False and (clave in Linf[i]) and opo in Lop[i]:
                            j = 1
                            z = True
                        if j == 2 and z == True:
                            j = 0
                            z = False
                            for x in range(x, y):
                                op.append(x)
                                if x == y-1:
                                    x = y
                if opo != '':
                    if opo == '1':
                        name = name+' 1eras'
                    elif opo == '2':
                        name = name+' 2das'
                    elif opo == '3':
                        name = name+' 3ras'
                    elif opo == '4':
                        name = name+' 4tas'
                    elif opo == '5':
                        name = name+' 5tas'
                    elif opo == '6':
                        name = name+' 6tas'
                if clave != '':
                    name = name+ ' '+clave
                # Se crea un nuevo DataFrame basado en las coordenadas obtenidas del original
                print(count)
                header3 = "HORARIO ALUMNOS - TOTAL DE ALUMNOS: " + str(count)
                df2 = df.iloc[op, :]
                writer = pd.ExcelWriter(name+'.xlsx', engine='xlsxwriter')
                if opo != '':
                    df2.to_excel(writer, startrow=7, startcol=0,
                                engine='xlsxwriter', index=False)
                else:
                    df.to_excel(writer, startrow=7, startcol=0,
                                engine='xlsxwriter', index=False)
                col_idx_info = df.columns.get_loc('Informacion')
                col_idx_rec = df.columns.get_loc('Frecuencia')
                worksheet = writer.sheets['Sheet1']
                worksheet.write(1, 0, header1)
                worksheet.write(2, 0, header2)
                worksheet.write(3, 0, header3)
                worksheet.write(4, 0, header4)
                worksheet.write(5, 0, header5)
                worksheet.set_column(col_idx_info, col_idx_info, 45)
                worksheet.set_column(col_idx_rec, col_idx_rec, 12)
                writer.close()
                #success()
                y = 0
        else:
            y = 0
            messagebox.showinfo(message = "Por favor, elija los archivos.", title = "Aviso")
            #wrongopt()
            return